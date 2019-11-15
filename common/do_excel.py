# -*- coding: utf-8 -*-
# @Time : 2019/11/13 9:31
# @Author : kira
# @Email : 262667641@qq.com
# @File : do_excel.py
# @Project : risk_project

from openpyxl import load_workbook
from common import project_path


class DoeExcel():
    def __init__(self, file_name=None, sheet_name=None):
        if file_name:
            self.file_name = file_name
            self.sheet_name = sheet_name
        else:
            self.file_name = r"..\data\risk_api.xlsx"
            self.sheet_name = "vip"
        self.wb = load_workbook(self.file_name)  # 获取 excel 句柄
        self.sheet = self.wb[self.sheet_name]  # 依据 sheet_name 获取 sheet
        self.max_row = self.sheet.max_row  # 获取 sheet 最大行
        self.max_column = self.sheet.max_column  # 获取 sheet 最大列

    def get_headers(self):
        '''
        获取 sheet 中的第一列标题栏的所有值
        :return:
        '''
        all_headers = []
        for i in range(1, self.max_column + 1):
            each_title = self.sheet.cell(1, i).value
            all_headers.append(each_title)
        return all_headers

    def do_excel(self):
        '''
        通过 title 定位单元格，获取所有测试数据
        :return:
        '''
        all_headers = self.get_headers()
        test_data = []
        for i in range(1, self.max_row):
            print(range(1, self.max_row))
            sub_data = {}
            for j in range(1, self.max_row + 1):
                print(j)
                sub_data[all_headers[j - 1]] = self.sheet.cell(i + 1, j).value
            test_data.append(sub_data)
        return test_data


if __name__ == '__main__':
    tt = DoeExcel(project_path.data_path, "info")
    print("\n", tt.do_excel())
