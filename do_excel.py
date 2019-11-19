# -*- coding: utf-8 -*-
# @Time : 2019/11/13 9:31
# @Author : kira
# @Email : 262667641@qq.com
# @File : do_excel.py
# @Project : risk_project

from openpyxl import load_workbook
from common import project_path
from common.get_conf_data import GetConfigData
import logging


class DoeExcel():

    def do_excel(self, file_name):
        '''
        通过 title 定位单元格，获取所有测试数据
        :return:
        '''
        mode = eval(GetConfigData().get_config_data(project_path.conf_path, "CASE", "sheet_list"))
        wb = load_workbook(file_name)
        test_data = []
        for key in mode:  # 遍历存在配置文件里面的字典，key==sheet_name
            sheet = wb[key]  # 获取所有 sheet
            max_row = sheet.max_row  # 获取最大行
            max_column = sheet.max_column  # 获取最大列
            fist_header = []  # 获取第一行标题所有值
            for i in range(1, max_column + 1):
                fist_header.append(sheet.cell(1, i).value)
            if mode[key] == "all":
                # 定位单元格
                for i in range(2, max_row + 1):
                    sub_data = {}  # 列表内的字典（也就是测试数据）
                    for k in range(1, max_column + 1):
                        sub_data[fist_header[k - 1]] = sheet.cell(i, k).value
                        sub_data["sheet_name"] = key  # 多添加一列值 sheet_name
                    test_data.append(sub_data)  # 将所有单元格 title 对应的值组成字典添加到列表中。
            else:
                for case_id in mode[key]:
                    sub_data = {}  # 列表内的字典（也就是测试数据）
                    for k in range(1, max_column + 1):
                        sub_data[fist_header[k - 1]] = sheet.cell(case_id + 1, k).value  # 将每一个 case_id 行的值与title 组成字典
                        sub_data["sheet_name"] = key  # 多添加一列 sheet_name
                    # print(sub_data)
                    test_data.append(sub_data)
        return test_data

    def write_back(self, file_name, sheet_name, i, respone_value, test_resulet):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i + 1, 9).value = respone_value
        sheet.cell(i + 1, 10).value = test_resulet
        wb.save(file_name)
        wb.close()

    def read_respone_result(self,file_name,sheet_name,i, j):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        respone_value = sheet.cell(i, j).value
        wb.close()
        return respone_value





if __name__ == '__main__':
    tt = DoeExcel()
    test = tt.do_excel(project_path.data_path)
    print(test)
    print(tt.write_back(project_path.data_path, ))
